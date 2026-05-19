import json
from pathlib import Path
from collections import defaultdict

from flask import Flask, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from checker import validator
from excel_to_json import convert_excel_to_json, create_template
from json_to_calendar import convert_json_to_calendar
from scheduler import schedule as build_schedule

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
for folder in (UPLOAD_DIR, OUTPUT_DIR):
    folder.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False

DAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
LESSON_HEADERS = ["день_недели", "время", "группа", "предмет", "тип", "преподаватель", "аудитория"]


def save_json(data, path: Path):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=4), encoding="utf-8")


def lessons_by_group(schedule_data):
    lessons = schedule_data.get("lessons", []) if isinstance(schedule_data, dict) else []
    groups = defaultdict(list)
    for lesson in lessons:
        raw_groups = str(lesson.get("группа", "")).split(",")
        for group in [g.strip() for g in raw_groups if g.strip()]:
            groups[group].append(lesson)
    return dict(sorted(groups.items(), key=lambda item: item[0]))


def sorted_lessons(lessons):
    day_index = {day: i for i, day in enumerate(DAYS)}
    return sorted(lessons, key=lambda x: (day_index.get(x.get("день_недели", ""), 99), str(x.get("время", ""))))


def autosize(sheet):
    for column_cells in sheet.columns:
        max_len = 10
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        sheet.column_dimensions[letter].width = min(max_len + 3, 42)


def write_schedule_sheet(workbook, title, lessons):
    sheet = workbook.create_sheet(title[:31])
    sheet.append(["День недели", "Время", "Группа", "Предмет", "Тип", "Преподаватель", "Аудитория"])
    fill = PatternFill("solid", fgColor="3657FF")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for lesson in sorted_lessons(lessons):
        sheet.append([
            lesson.get("день_недели", ""), lesson.get("время", ""), lesson.get("группа", ""),
            lesson.get("предмет", ""), lesson.get("тип", ""), lesson.get("преподаватель", ""), lesson.get("аудитория", ""),
        ])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    autosize(sheet)


def export_schedule_excel(schedule_data, output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    lessons = schedule_data.get("lessons", [])
    write_schedule_sheet(workbook, "Общее расписание", lessons)
    for group, group_lessons in lessons_by_group(schedule_data).items():
        write_schedule_sheet(workbook, f"Группа {group}", group_lessons)
    workbook.save(output_path)


def export_schedule_pdf(schedule_data, output_path):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "Helvetica"
    for candidate in ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/Library/Fonts/Arial Unicode.ttf", "C:/Windows/Fonts/arial.ttf"]:
        if Path(candidate).exists():
            pdfmetrics.registerFont(TTFont("LocalSans", candidate))
            font_name = "LocalSans"
            break

    doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name
    story = []

    def add_table(title, lessons):
        story.append(Paragraph(title, styles["Title"]))
        story.append(Spacer(1, 10))
        rows = [["День недели", "Время", "Группа", "Предмет", "Тип", "Преподаватель", "Аудитория"]]
        for lesson in sorted_lessons(lessons):
            rows.append([str(lesson.get(key, "")) for key in LESSON_HEADERS])
        table = Table(rows, repeatRows=1, colWidths=[78, 70, 72, 180, 70, 170, 70])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(table)

    add_table("Общее расписание", schedule_data.get("lessons", []))
    for group, group_lessons in lessons_by_group(schedule_data).items():
        story.append(PageBreak())
        add_table(f"Группа {group}", group_lessons)
    doc.build(story)


@app.get("/")
def index():
    return render_template("index.html")


@app.get("/download/template")
def download_template():
    template_path = OUTPUT_DIR / "schedule_data_template.xlsx"
    if not template_path.exists():
        create_template(template_path)
    return send_file(template_path, as_attachment=True, download_name="schedule_data_template.xlsx")


@app.post("/api/upload-excel")
def upload_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Файл не загружен"}), 400
    filename = secure_filename(file.filename or "schedule_data.xlsx")
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return jsonify({"error": "Загрузите Excel-файл .xlsx по шаблону."}), 400
    input_path = UPLOAD_DIR / filename
    output_path = OUTPUT_DIR / f"{Path(filename).stem}.json"
    file.save(input_path)
    try:
        data = convert_excel_to_json(input_path, output_path)
    except Exception as exc:
        return jsonify({"error": f"Файл не соответствует шаблону: {exc}. Скачайте шаблон, заполните его и загрузите снова."}), 400
    return jsonify({"data": data, "message": "Файл загружен"})


@app.post("/api/generate")
def generate():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Нет данных для создания расписания"}), 400
    try:
        result = build_schedule(data)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    save_json(result, OUTPUT_DIR / "schedule.json")
    return jsonify(result)


@app.post("/api/validate")
def validate_schedule():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Нет расписания для проверки"}), 400
    try:
        errors = validator(data)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"errors": errors, "count": len(errors)})


@app.post("/api/calendar")
def calendar():
    payload = request.get_json(silent=True) or {}
    schedule_data = payload.get("schedule")
    if not schedule_data:
        return jsonify({"error": "Нет расписания для экспорта"}), 400
    input_path = OUTPUT_DIR / "schedule_for_calendar.json"
    output_path = OUTPUT_DIR / "schedule.ics"
    save_json(schedule_data, input_path)
    try:
        convert_json_to_calendar(input_path, output_path, week_start=payload.get("week_start") or None, timezone=payload.get("timezone") or "Asia/Novosibirsk", repeat_weeks=int(payload.get("repeat_weeks") or 16))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return send_file(output_path, as_attachment=True, download_name="schedule.ics", mimetype="text/calendar")


@app.post("/api/excel")
def excel_export():
    payload = request.get_json(silent=True) or {}
    schedule_data = payload.get("schedule")
    if not schedule_data:
        return jsonify({"error": "Нет расписания для экспорта"}), 400
    output_path = OUTPUT_DIR / "schedule_export.xlsx"
    export_schedule_excel(schedule_data, output_path)
    return send_file(output_path, as_attachment=True, download_name="schedule_export.xlsx")


@app.post("/api/pdf")
def pdf_export():
    payload = request.get_json(silent=True) or {}
    schedule_data = payload.get("schedule")
    if not schedule_data:
        return jsonify({"error": "Нет расписания для экспорта"}), 400
    output_path = OUTPUT_DIR / "schedule_export.pdf"
    try:
        export_schedule_pdf(schedule_data, output_path)
    except Exception as exc:
        return jsonify({"error": f"Не удалось создать PDF: {exc}"}), 400
    return send_file(output_path, as_attachment=True, download_name="schedule_export.pdf", mimetype="application/pdf")


if __name__ == "__main__":
    app.run(debug=True)
